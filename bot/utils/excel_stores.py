from datetime import datetime
from io import BytesIO

from openpyxl import Workbook

from bot.db.models import Store
from bot.utils.store_flow import TZ_TASHKENT, fmt_store_date


def stores_to_xlsx_bytes(stores: list[Store]) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Magazinlar"
    ws.append(
        [
            "ID",
            "Nomi",
            "Telefon",
            "Manzil",
            "Hisobot sanasi",
            "Oylik (so'm)",
            "Qarz (so'm)",
            "Tok kW (joriy)",
            "Tok qarz / oxirgi +kW",
            "Yaratilgan",
        ]
    )
    for s in stores:
        created = "—"
        if s.created_at:
            created = s.created_at.astimezone(TZ_TASHKENT).strftime("%d.%m.%Y %H:%M")
        ws.append(
            [
                s.id,
                s.name,
                s.owner_phone or "",
                s.address or "",
                fmt_store_date(s.store_date),
                s.monthly_amount if s.monthly_amount is not None else "",
                int(s.debt_balance or 0),
                s.electricity_kw if s.electricity_kw is not None else "",
                int(s.debt_tok or 0),
                created,
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _fmt_ts(ca: datetime) -> str:
    if ca.tzinfo is None:
        ca = ca.replace(tzinfo=TZ_TASHKENT)
    return ca.astimezone(TZ_TASHKENT).strftime("%d.%m.%Y %H:%M")


def admin_report_xlsx_bytes(
    stores: list[Store],
    payment_rows: list[
        tuple[int, int, str, int, int, datetime, int]
    ],  # id, sid, name, amount, debt_after, created_at, admin_id
    electricity_rows: list[
        tuple[int, int, str, datetime, datetime, int, int, int, datetime]
    ],  # log_id, sid, name, p_from, p_to, rb, ra, delta, created_at
) -> bytes:
    """Bitta fayl: magazinlar + qarzdan ayirishlar + tok tarixi (kW)."""
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Magazinlar"
    ws1.append(
        [
            "ID",
            "Nomi",
            "Telefon",
            "Manzil",
            "Hisobot sanasi",
            "Oylik (so'm)",
            "Qarz (so'm)",
            "Tok kW (joriy)",
            "Tok qarz / oxirgi +kW",
            "Yaratilgan",
        ]
    )
    for s in stores:
        created = "—"
        if s.created_at:
            created = s.created_at.astimezone(TZ_TASHKENT).strftime("%d.%m.%Y %H:%M")
        ws1.append(
            [
                s.id,
                s.name,
                s.owner_phone or "",
                s.address or "",
                fmt_store_date(s.store_date),
                s.monthly_amount if s.monthly_amount is not None else "",
                int(s.debt_balance or 0),
                s.electricity_kw if s.electricity_kw is not None else "",
                int(s.debt_tok or 0),
                created,
            ]
        )

    ws2 = wb.create_sheet("Qarzdan ayirishlar")
    ws2.append(
        [
            "ID",
            "Magazin ID",
            "Magazin nomi",
            "Sana vaqt",
            "Ayirilgan summa (so'm)",
            "Qarzdan keyin (so'm)",
            "Admin Telegram ID",
        ]
    )
    for row in payment_rows:
        pid, sid, name, amount, debt_after, created_at, admin_id = row
        ts = _fmt_ts(created_at)
        ws2.append([pid, sid, name, ts, amount, debt_after, admin_id])

    ws3 = wb.create_sheet("Tok tarixi (kW)")
    ws3.append(
        [
            "ID",
            "Magazin ID",
            "Magazin nomi",
            "Davr boshlanishi",
            "Davr tugashi",
            "Eski ko'rsatkich (kW)",
            "Yangi ko'rsatkich (kW)",
            "Iste'mol (+kW)",
            "Yozilgan vaqt",
        ]
    )
    for row in electricity_rows:
        lid, sid, name, p_from, p_to, rb, ra, delta, created_at = row
        ws3.append(
            [
                lid,
                sid,
                name,
                _fmt_ts(p_from),
                _fmt_ts(p_to),
                rb,
                ra,
                delta,
                _fmt_ts(created_at),
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
